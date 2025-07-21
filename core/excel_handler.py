import pandas as pd
import os
from openpyxl.styles import Font


class ExcelHandler:
    """
    Handles Excel file operations like reading, writing, and formatting.
    """

    @staticmethod
    def read_excel(file_path):
        """
        Read an Excel file and return a pandas DataFrame.

        Args:
            file_path (str): Path to the Excel file

        Returns:
            pandas.DataFrame: The data from the Excel file
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        return pd.read_excel(file_path)

    @staticmethod
    def create_pivot_table(df, values, index, aggfunc='sum'):
        """
        Create a pivot table from a DataFrame.

        Args:
            df (pandas.DataFrame): The source DataFrame
            values (str or list): Column(s) to aggregate
            index (list): Columns to group by
            aggfunc (str or function): Aggregation function

        Returns:
            pandas.DataFrame: The resulting pivot table
        """
        pivot_df = df.pivot_table(
            values=values,
            index=index,
            aggfunc=aggfunc
        ).reset_index()

        return pivot_df

    # @staticmethod
    # def write_excel(df, output_file, sheet_name='Sheet1'):
    #     """
    #     Write a DataFrame to an Excel file with formatting.
    #
    #     Args:
    #         df (pandas.DataFrame): The data to write
    #         output_file (str): Path where the output file will be saved
    #         sheet_name (str): Name of the worksheet
    #     """
    #     # Create directory if it doesn't exist
    #     output_dir = os.path.dirname(output_file)
    #     if output_dir and not os.path.exists(output_dir):
    #         os.makedirs(output_dir)
    #
    #     with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    #         df.to_excel(writer, index=False, sheet_name=sheet_name)
    #
    #         # Get the workbook and worksheet
    #         workbook = writer.book
    #         worksheet = writer.sheets[sheet_name]
    #
    #         # Apply formatting for resource rows (non-indented)
    #         for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1)):
    #             cell_value = str(row[0].value) if row[0].value else ""
    #             if not cell_value.startswith('    '):
    #                 # Make resource rows bold
    #                 for cell in row:
    #                     cell.font = Font(bold=True)
    #
    #     return output_file
    @staticmethod
    def write_excel(df, output_file, sheet_name='Sheet1'):
        """
        Write a DataFrame to an Excel file with formatting.

        Args:
            df (pandas.DataFrame): The data to write
            output_file (str): Path where the output file will be saved
            sheet_name (str): Name of the worksheet
        """
        # Create directory if it doesn't exist
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Find the Ecart column index
            ecart_col_idx = None
            for idx, col_name in enumerate(df.columns):
                if col_name == 'Ecart':
                    ecart_col_idx = idx
                    break

            # Apply formatting for resource rows (non-indented) and Ecart column
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1)):
                cell_value = str(row[0].value) if row[0].value else ""

                # Make resource rows bold
                if not cell_value.startswith('    '):
                    for cell in row:
                        cell.font = Font(bold=True)

                # Apply conditional formatting to Ecart column
                if ecart_col_idx is not None and not cell_value.startswith('    '):
                    # Skip resource rows for Ecart formatting (they don't have Ecart values)
                    continue

                if ecart_col_idx is not None:
                    ecart_cell = row[ecart_col_idx]
                    if ecart_cell.value is not None and isinstance(ecart_cell.value, (int, float)):
                        from openpyxl.styles import PatternFill
                        if ecart_cell.value > 0:  # Positive value
                            # Light green color (vert accentuation6 plus clair 60%)
                            ecart_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                        elif ecart_cell.value < 0:  # Negative value
                            # Light red color (same grade but red)
                            ecart_cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        return output_file

    @staticmethod
    def write_multiple_sheets(dfs_dict, output_file):
        """
        Write multiple DataFrames to an Excel file, each in its own sheet, with formatting.

        Args:
            dfs_dict (dict): {sheet_name: DataFrame}
            output_file (str): Path where the output file will be saved
        """
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df in dfs_dict.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                # Find the Ecart column index
                ecart_col_idx = None
                for idx, col_name in enumerate(df.columns):
                    if col_name == 'Ecart':
                        ecart_col_idx = idx
                        break
                # Apply formatting for resource rows (non-indented) and Ecart column
                for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1)):
                    cell_value = str(row[0].value) if row[0].value else ""
                    # Make resource rows bold
                    if not cell_value.startswith('    '):
                        for cell in row:
                            cell.font = Font(bold=True)
                    # Apply conditional formatting to Ecart column
                    if ecart_col_idx is not None and not cell_value.startswith('    '):
                        continue
                    if ecart_col_idx is not None:
                        ecart_cell = row[ecart_col_idx]
                        if ecart_cell.value is not None and isinstance(ecart_cell.value, (int, float)):
                            from openpyxl.styles import PatternFill
                            if ecart_cell.value > 0:
                                ecart_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                            elif ecart_cell.value < 0:
                                ecart_cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        return output_file

    @staticmethod
    def open_file(file_path):
        """
        Open a file with the default application.

        Args:
            file_path (str): Path to the file to open

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            if os.name == 'nt':  # Windows
                os.startfile(file_path)
            elif os.name == 'posix':  # macOS or Linux
                import subprocess
                try:
                    subprocess.call(['open', file_path])  # macOS
                except:
                    subprocess.call(['xdg-open', file_path])  # Linux
            return True
        except Exception as e:
            print(f"Could not open file: {e}")
            return False

    @staticmethod
    def add_graphs_sheet(output_file, sheet_name, figures):
        """
        Add a new sheet to the Excel file and insert matplotlib figures as images, aligned horizontally.

        Args:
            output_file (str): Path to the Excel file to modify
            sheet_name (str): Name of the sheet to add
            figures (list): List of matplotlib Figure objects
        """
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        import tempfile
        import os

        # Load the workbook
        wb = load_workbook(output_file)
        # Remove existing sheet if present
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

        temp_files = []
        col_letters = ['A', 'K', 'U', 'AE', 'AO', 'AY']  # Add more if needed
        for idx, fig in enumerate(figures):
            tmpfile = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            fig.savefig(tmpfile.name, bbox_inches='tight')
            tmpfile.flush()
            img = XLImage(tmpfile.name)
            col = col_letters[idx] if idx < len(col_letters) else f'A{idx*10+1}'
            ws.add_image(img, f'{col}1')
            temp_files.append(tmpfile.name)
            tmpfile.close()
        wb.save(output_file)
        # Now delete temp files
        for fname in temp_files:
            try:
                os.unlink(fname)
            except Exception:
                pass