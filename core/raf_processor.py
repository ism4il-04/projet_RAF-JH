import pandas as pd
import calendar
from config.raf_rules import get_raf
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference


class RAFProcessor:
    """
    Processes RAF (Reste À Faire) calculations and generates RAF reports.
    """

    @staticmethod
    def calculate_raf(df):
        """
        Calculate RAF values for each deployment based on connection level and project phase.

        Args:
            df (pandas.DataFrame): The deployments DataFrame with 'Niveau de connexion' and 'Phase du projet'

        Returns:
            pandas.DataFrame: DataFrame with added RAF column
        """
        result_df = df.copy()

        # Add RAF column with NaN values
        result_df['RAF'] = None

        # Calculate RAF for each row
        for idx, row in result_df.iterrows():
            connection_level = row.get('Niveau de connexion')
            project_phase = row.get('Phase du projet')

            if connection_level and project_phase:
                raf_value = get_raf(connection_level, project_phase)
                result_df.at[idx, 'RAF'] = raf_value

        return result_df

    @staticmethod
    def add_raf_to_workbook(workbook, deployments_df):
        """
        Add RAF column to the active sheet of an existing workbook.

        Args:
            workbook: The openpyxl workbook to modify
            deployments_df (pandas.DataFrame): DataFrame with RAF values

        Returns:
            openpyxl.Workbook: The modified workbook
        """
        # Get the active sheet
        sheet = workbook.active

        # Find the last column and add RAF header
        last_col = sheet.max_column + 1
        sheet.cell(row=1, column=last_col, value="RAF")

        # Add RAF values
        row_index = 2  # Start from row 2 (after header)
        for raf_value in deployments_df['RAF'].values:
            sheet.cell(row=row_index, column=last_col, value=raf_value)
            row_index += 1

        return workbook

    @staticmethod
    def create_raf_summary_sheet(workbook, deployments_df):
        """
        Create a RAF summary sheet in the workbook with weekly and monthly breakdowns.

        Args:
            workbook: The openpyxl workbook to modify
            deployments_df (pandas.DataFrame): DataFrame with RAF values and Date de MEP

        Returns:
            openpyxl.Workbook: The modified workbook
        """
        # Add a new sheet for RAF summary
        if "RAF Summary" in workbook.sheetnames:
            # Remove existing sheet if it exists
            del workbook["RAF Summary"]

        raf_sheet = workbook.create_sheet(title="RAF Summary")

        # Calculate weekly RAF data
        # Ensure Date de MEP is datetime
        deployments_df = deployments_df.copy()
        deployments_df['Date de MEP'] = pd.to_datetime(deployments_df['Date de MEP'], errors='coerce')

        # Create Year-Month groups first
        year_month_groups = {}
        for year in deployments_df['Date de MEP'].dt.year.dropna().unique():
            year_data = deployments_df[deployments_df['Date de MEP'].dt.year == year]
            months_in_year = {}

            for month in year_data['Date de MEP'].dt.month.unique():
                month_name = calendar.month_name[month]
                month_data = year_data[year_data['Date de MEP'].dt.month == month]

                # Calculate total RAF for the month
                month_raf = month_data['RAF'].sum()

                # Skip months with zero RAF
                if month_raf == 0:
                    continue

                # Group by week of month
                weeks_data = []
                # Get week numbers and find min/max days for each week
                week_groups = {}

                for _, row in month_data.iterrows():
                    if pd.notna(row['Date de MEP']):
                        week_num = row['Date de MEP'].isocalendar()[1]
                        day = row['Date de MEP'].day

                        if week_num not in week_groups:
                            week_groups[week_num] = {'days': [], 'raf': 0}

                        week_groups[week_num]['days'].append(day)
                        if pd.notna(row['RAF']):
                            week_groups[week_num]['raf'] += row['RAF']

                # Process each week to create formatted week entries
                for week_num, week_data in week_groups.items():
                    raf_value = week_data['raf']

                    # Only include weeks with non-zero RAF
                    if raf_value > 0:
                        min_day = min(week_data['days']) if week_data['days'] else 0
                        max_day = max(week_data['days']) if week_data['days'] else 0

                        # Format as "Week X (03 to 07)"
                        week_label = f"Week {week_num}"
                        if min_day and max_day:
                            if min_day == max_day:
                                day_range = f"({min_day:02d})"
                            else:
                                day_range = f"({min_day:02d} to {max_day:02d})"
                            week_label = f"{week_label} {day_range}"

                        weeks_data.append((week_label, raf_value))

                # Only add month if it has weeks with non-zero RAF
                if weeks_data:
                    months_in_year[month_name] = {
                        'total_raf': month_raf,
                        'weeks': weeks_data
                    }

            # Only add year if it has months with data
            if months_in_year:
                year_month_groups[year] = months_in_year

        # Create a stylish report
        # Add title
        raf_sheet.merge_cells('A1:E1')
        title_cell = raf_sheet.cell(row=1, column=1, value="RAF Summary Report")
        title_cell.font = Font(size=16, bold=True, color="000080")  # Navy color
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        total_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")  # Light blue
        year_font = Font(size=14, bold=True, color="FFFFFF")
        header_font = Font(size=12, bold=True, color="FFFFFF")
        month_font = Font(size=12, bold=True)
        week_font = Font(size=11)

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        medium_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Set column widths
        raf_sheet.column_dimensions['A'].width = 35  # Period column - wider for week ranges
        raf_sheet.column_dimensions['B'].width = 15  # RAF Value column

        # Start row for content
        current_row = 3

        # Table headers
        raf_sheet.cell(row=current_row, column=1, value="Period").font = header_font
        raf_sheet.cell(row=current_row, column=2, value="RAF Value").font = header_font
        for col in range(1, 3):
            raf_sheet.cell(row=current_row, column=col).fill = header_fill
            raf_sheet.cell(row=current_row, column=col).border = medium_border
            raf_sheet.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        current_row += 1

        # Add data by year, month, and week in a consolidated table
        for year, months in sorted(year_month_groups.items()):
            # Year row
            year_cell = raf_sheet.cell(row=current_row, column=1, value=f"Year {int(year)}")
            year_cell.font = year_font
            year_cell.fill = header_fill
            year_cell.alignment = Alignment(horizontal='left')
            year_cell.border = thin_border

            raf_sheet.cell(row=current_row, column=2).border = thin_border
            raf_sheet.cell(row=current_row, column=2).fill = header_fill

            current_row += 1

            # Add months and weeks
            for month_name, month_data in sorted(months.items(),
                                                 key=lambda x: list(calendar.month_name).index(x[0]) if x[
                                                                                                            0] in calendar.month_name else 0):
                # Month row
                month_cell = raf_sheet.cell(row=current_row, column=1, value=month_name)
                raf_cell = raf_sheet.cell(row=current_row, column=2, value=month_data['total_raf'])

                month_cell.font = month_font
                raf_cell.font = month_font
                month_cell.alignment = Alignment(horizontal='left')
                raf_cell.alignment = Alignment(horizontal='center')
                month_cell.border = thin_border
                raf_cell.border = thin_border
                month_cell.fill = total_fill
                raf_cell.fill = total_fill

                current_row += 1

                # Week rows - only if there are non-zero weeks
                for week_label, week_raf in sorted(month_data['weeks']):
                    # Skip weeks with zero RAF
                    if week_raf == 0:
                        continue

                    week_cell = raf_sheet.cell(row=current_row, column=1,
                                               value=f"  • {week_label}")  # Indent with bullet
                    raf_cell = raf_sheet.cell(row=current_row, column=2, value=week_raf)

                    week_cell.font = week_font
                    raf_cell.font = week_font
                    week_cell.alignment = Alignment(horizontal='left')
                    raf_cell.alignment = Alignment(horizontal='center')
                    week_cell.border = thin_border
                    raf_cell.border = thin_border

                    current_row += 1

        return workbook
